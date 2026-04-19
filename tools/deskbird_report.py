"""
deskbird_report.py — Local attendance comparison script.

Reads a JSON file of deskbird booking data, compares against scheduled
office days (from Supabase), and emails an XLSX report.

Usage:
    python deskbird_report.py deskbird_data.json
    python deskbird_report.py --from-supabase [--start-date YYYY-MM-DD] [--days 7]

Input JSON format:
    [
        {"date": "2026-04-13", "people": ["Alice Smith", "Bob Jones"]},
        {"date": "2026-04-14", "people": ["Alice Smith"]}
    ]

Config: reads from .env in the same directory as this script.
"""

import json
import os
import re
import smtplib
import sys
from collections import defaultdict
from datetime import date, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import create_client


# ── Config ─────────────────────────────────────────────────────────────────────

def load_env():
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        os.environ.setdefault(key.strip(), val.strip())


def get_env(key):
    val = os.environ.get(key, "").strip()
    if not val:
        raise SystemExit(f"Missing required config: {key}  (set it in tools/.env)")
    return val


# ── Name normalisation (same logic as the old attendance.py) ──────────────────

def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower()).strip()


def normalize_first_name(value: str) -> str:
    first = (value or "").strip().split()[0] if (value or "").strip() else ""
    return normalize_name(first)


def group_people_by_first_name(people_rows):
    grouped = defaultdict(list)
    for row in people_rows:
        grouped[normalize_first_name(row["name"])].append(row["name"])
    return grouped


# ── Schedule helper ────────────────────────────────────────────────────────────

def active_schedule_rows(rows, week_start_str: str) -> list:
    filtered = [
        r for r in rows
        if (r.get("valid_from") or "2000-01-01") <= week_start_str
        and (r.get("valid_until") is None or r["valid_until"] >= week_start_str)
    ]
    if not filtered:
        return []
    latest_version = max(r.get("valid_from") or "2000-01-01" for r in filtered)
    version_rows = [r for r in filtered if (r.get("valid_from") or "2000-01-01") == latest_version]
    result = {}
    for row in sorted(version_rows, key=lambda r: r.get("valid_from") or "2000-01-01"):
        key = (row.get("person_id", ""), row["day_of_week"])
        result[key] = row
    return list(result.values())


# ── Core comparison logic ──────────────────────────────────────────────────────

def compute_report(booking_days: list[dict]) -> list[dict]:
    """
    booking_days: [{"date": "YYYY-MM-DD", "people": ["Name", ...]}, ...]
    Returns per-day comparison dicts.
    """
    load_env()
    sb = create_client(get_env("SUPABASE_URL"), get_env("SUPABASE_SERVICE_ROLE_KEY"))

    target_day_strings = sorted({d["date"] for d in booking_days})

    people_rows = sb.table("people").select("id, name").eq("active", True).order("name").execute().data
    schedule_rows = (
        sb.table("person_schedule")
        .select("person_id, day_of_week, hours, location, valid_from, valid_until")
        .execute()
        .data
    )
    absence_rows = (
        sb.table("absences")
        .select("person_id, date")
        .gte("date", target_day_strings[0])
        .lte("date", target_day_strings[-1])
        .execute()
        .data
    )

    people_by_id = {row["id"]: row["name"] for row in people_rows}
    duplicate_first_name_groups = {
        fn: sorted(names)
        for fn, names in group_people_by_first_name(people_rows).items()
        if fn and len(names) > 1
    }
    known_match_keys = set()
    for person_name in people_by_id.values():
        first_name_key = normalize_first_name(person_name)
        if first_name_key in duplicate_first_name_groups:
            known_match_keys.add(normalize_name(person_name))
        else:
            known_match_keys.add(first_name_key)
    absent_by_date = defaultdict(set)
    for row in absence_rows:
        absent_by_date[row["date"]].add(row["person_id"])

    actual_by_date = {d["date"]: d["people"] for d in booking_days}
    actual_match_keys_by_date = defaultdict(set)
    for d in booking_days:
        for person in d["people"]:
            actual_match_keys_by_date[d["date"]].add(normalize_first_name(person))
            actual_match_keys_by_date[d["date"]].add(normalize_name(person))

    # Group schedule rows per person
    sched_by_pid = defaultdict(list)
    for r in schedule_rows:
        sched_by_pid[r["person_id"]].append(r)

    results = []
    for day_str in sorted(target_day_strings):
        d_obj = date.fromisoformat(day_str)
        dow = d_obj.weekday() + 1  # 1=Mon…5=Fri

        active_for_day = []
        for pid_rows in sched_by_pid.values():
            active_for_day.extend(active_schedule_rows(pid_rows, day_str))

        by_person = defaultdict(dict)
        for row in active_for_day:
            by_person[row["person_id"]][row["day_of_week"]] = {
                "hours": float(row.get("hours") or 0),
                "location": (row.get("location") or "office").lower(),
            }

        expected = []
        missing = []
        expected_first = set()
        for person_id, person_name in people_by_id.items():
            sched = by_person.get(person_id, {}).get(dow)
            if not sched or sched["hours"] <= 0 or sched["location"] != "office":
                continue
            if person_id in absent_by_date[day_str]:
                continue
            fn_key = normalize_first_name(person_name)
            use_full = fn_key in duplicate_first_name_groups
            norm_key = normalize_name(person_name) if use_full else fn_key
            expected.append(person_name)
            expected_first.add(norm_key)
            if norm_key not in actual_match_keys_by_date[day_str]:
                missing.append(person_name)

        actual = sorted(actual_by_date.get(day_str, []))
        unexpected = sorted(
            p for p in actual
            if (
                (normalize_name(p) if normalize_first_name(p) in duplicate_first_name_groups else normalize_first_name(p)) in known_match_keys
                and (normalize_name(p) if normalize_first_name(p) in duplicate_first_name_groups else normalize_first_name(p)) not in expected_first
            )
        )
        results.append({
            "date": day_str,
            "weekday": d_obj.strftime("%a"),
            "expected_office": sorted(expected),
            "actual_deskbird": actual,
            "missing_bookings": sorted(missing),
            "unexpected_bookings": unexpected,
            "absent_people": sorted(
                people_by_id[pid]
                for pid in absent_by_date[day_str]
                if pid in people_by_id
            ),
        })
    return results


def workdays_from(start_date: date, days: int) -> list[date]:
    current = start_date
    output: list[date] = []
    while len(output) < days:
        if current.weekday() < 5:
            output.append(current)
        current += timedelta(days=1)
    return output


def load_booking_days_from_supabase(start_date_str: str | None = None, days: int = 7) -> list[dict]:
    load_env()
    sb = create_client(get_env("SUPABASE_URL"), get_env("SUPABASE_SERVICE_ROLE_KEY"))

    start = date.fromisoformat(start_date_str) if start_date_str else date.today()
    target_days = workdays_from(start, days)
    target_strings = [str(d) for d in target_days]

    latest = (
        sb.table("deskbird_sync_runs")
        .select("*")
        .order("fetched_at", desc=True)
        .limit(1)
        .execute()
        .data
    )
    if not latest:
        raise SystemExit("No deskbird sync run found in Supabase")
    run = latest[0]

    bookings = (
        sb.table("deskbird_attendance_bookings")
        .select("booking_date, person_name")
        .eq("sync_run_id", run["id"])
        .in_("booking_date", target_strings)
        .execute()
        .data
    )

    booking_days_map = defaultdict(list)
    for row in bookings:
        booking_days_map[row["booking_date"]].append(row["person_name"])

    return [
        {"date": day_str, "people": sorted(booking_days_map.get(day_str, []))}
        for day_str in target_strings
    ]


# ── Excel generation ───────────────────────────────────────────────────────────

def build_xlsx(report: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    TITLE_FILL = PatternFill("solid", fgColor="1E3A8A")
    HDR_FILL = PatternFill("solid", fgColor="334155")
    MISS_FILL = PatternFill("solid", fgColor="FEE2E2")
    UNEX_FILL = PatternFill("solid", fgColor="FEF3C7")
    OK_FILL = PatternFill("solid", fgColor="ECFDF5")
    DATE_FILL = PatternFill("solid", fgColor="EFF6FF")
    META_FILL = PatternFill("solid", fgColor="F8FAFC")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, color="FFFFFF", size=13)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    bold_font = Font(bold=True, size=10)
    small_font = Font(size=9)
    small_bold = Font(bold=True, size=9)

    total_missing = sum(len(day["missing_bookings"]) for day in report)
    total_unexpected = sum(len(day["unexpected_bookings"]) for day in report)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws["A1"] = f"Deskbird Attendance Report  |  {report[0]['date']} to {report[-1]['date']}"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    summary = [
        ("Generated range", f"{report[0]['date']} to {report[-1]['date']}"),
        ("Missing bookings", str(total_missing)),
        ("Unexpected bookings", str(total_unexpected)),
    ]
    for col, (label, value) in enumerate(summary, start=1):
        label_cell = ws.cell(row=2, column=(col - 1) * 2 + 1)
        value_cell = ws.cell(row=2, column=(col - 1) * 2 + 2)
        label_cell.value = label
        value_cell.value = value
        for cell in (label_cell, value_cell):
            cell.fill = META_FILL
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.font = small_bold
        value_cell.font = small_font
    ws.row_dimensions[2].height = 18

    headers = ["Date", "Day", "Expected (office)", "Actual (Deskbird)",
               "Missing bookings", "Unexpected bookings", "Absent"]
    header_row = 4
    ws.append([])
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HDR_FILL
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 20

    for ri, day in enumerate(report, start=header_row + 1):
        has_missing    = bool(day["missing_bookings"])
        has_unexpected = bool(day["unexpected_bookings"])
        row_fill = MISS_FILL if has_missing else (UNEX_FILL if has_unexpected else OK_FILL)

        def nl(lst):
            return "\n".join(lst) if lst else "—"

        values = [
            day["date"],
            day["weekday"],
            nl(day["expected_office"]),
            nl(day["actual_deskbird"]),
            nl(day["missing_bookings"]) if day["missing_bookings"] else "—",
            nl(day["unexpected_bookings"]) if day["unexpected_bookings"] else "—",
            nl(day["absent_people"]) if day["absent_people"] else "—",
        ]
        ws.append(values)
        max_lines = max(max(len(v.split("\n")) for v in values if isinstance(v, str)), 1)
        ws.row_dimensions[ri].height = max(18, max_lines * 14)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=col)
            cell.fill = DATE_FILL if col <= 2 else row_fill
            cell.font = bold_font if col <= 2 else small_font
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if col <= 2 else "left")
            cell.border = border

    col_widths = [13, 8, 26, 26, 24, 24, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Email ──────────────────────────────────────────────────────────────────────

def send_email(xlsx_bytes: bytes, filename: str, report: list[dict]):
    load_env()
    smtp_host = get_env("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = get_env("SMTP_USER")
    smtp_pass = get_env("SMTP_PASSWORD")
    email_to  = get_env("EMAIL_TO")
    email_from = os.environ.get("EMAIL_FROM", smtp_user).strip()

    total_missing = sum(len(d["missing_bookings"]) for d in report)
    subject = f"Deskbird Attendance Report ({report[0]['date']} – {report[-1]['date']}) — {total_missing} missing"

    body_lines = [f"Deskbird attendance report: {report[0]['date']} to {report[-1]['date']}\n"]
    for day in report:
        body_lines.append(f"{day['date']} ({day['weekday']})")
        body_lines.append(f"  Expected:    {', '.join(day['expected_office']) or '—'}")
        body_lines.append(f"  Actual:      {', '.join(day['actual_deskbird']) or '—'}")
        if day["missing_bookings"]:
            body_lines.append(f"  MISSING:     {', '.join(day['missing_bookings'])}")
        if day["unexpected_bookings"]:
            body_lines.append(f"  Unexpected:  {', '.join(day['unexpected_bookings'])}")
        body_lines.append("")

    msg = MIMEMultipart()
    msg["From"] = email_from
    msg["To"] = email_to
    msg["Subject"] = subject
    msg.attach(MIMEText("\n".join(body_lines), "plain"))

    attachment = MIMEApplication(xlsx_bytes, _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    attachment.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(attachment)

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.ehlo()
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(email_from, email_to.split(","), msg.as_string())

    print(f"Email sent to {email_to}")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    from_supabase = False
    start_date_arg = None
    days = 7

    i = 0
    positional = []
    while i < len(args):
        arg = args[i]
        if arg == "--from-supabase":
            from_supabase = True
            i += 1
        elif arg == "--start-date":
            start_date_arg = args[i + 1]
            i += 2
        elif arg == "--days":
            days = int(args[i + 1])
            i += 2
        else:
            positional.append(arg)
            i += 1

    if from_supabase:
        booking_days = load_booking_days_from_supabase(start_date_arg, days)
    else:
        if not positional:
            print("Usage: python deskbird_report.py <deskbird_data.json>")
            print("   or: python deskbird_report.py --from-supabase [--start-date YYYY-MM-DD] [--days 7]")
            sys.exit(1)
        input_path = Path(positional[0])
        if not input_path.exists():
            print(f"File not found: {input_path}")
            sys.exit(1)
        booking_days = json.loads(input_path.read_text())

    if not booking_days:
        print("No booking data found for the requested range.")
        sys.exit(1)

    print(f"Processing {len(booking_days)} day(s)…")
    report = compute_report(booking_days)

    xlsx_bytes = build_xlsx(report)
    start, end = report[0]["date"], report[-1]["date"]
    filename = f"deskbird_report_{start}_to_{end}.xlsx" if start != end else f"deskbird_report_{start}.xlsx"

    # Save locally
    out_path = Path(filename)
    out_path.write_bytes(xlsx_bytes)
    print(f"Report saved: {out_path}")

    # Send email if configured
    load_env()
    if os.environ.get("EMAIL_TO", "").strip():
        print("Sending email…")
        send_email(xlsx_bytes, filename, report)
    else:
        print("EMAIL_TO not set — skipping email. Report saved locally only.")


if __name__ == "__main__":
    main()
