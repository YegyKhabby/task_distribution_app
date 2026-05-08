from collections import defaultdict
from fastapi import APIRouter, Query
from fastapi.responses import Response, StreamingResponse
from database import supabase
from datetime import date, timedelta
import calendar as cal_module
from io import BytesIO
from models import ActualHoursCreate, ActualHoursUpdate, CopyWeekRequest, ActualLocationUpsert
from routers.calendar import (
    distribute_week,
    get_mondays_in_month,
    holiday_dates_from_absence_rows,
    holiday_dows_for_week,
    _overlay_reallocations_on_tasks,
)
from utils.versioned import active_schedule_rows, active_distribution_rows
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/actual", tags=["actual"])


def _sf(hex6: str):
    return PatternFill("solid", fgColor=hex6)


def _lighten(hex6: str, factor: float = 0.91) -> str:
    r = int(hex6[0:2], 16)
    g = int(hex6[2:4], 16)
    b = int(hex6[4:6], 16)
    r2 = int(r + (255 - r) * factor)
    g2 = int(g + (255 - g) * factor)
    b2 = int(b + (255 - b) * factor)
    return f"{r2:02X}{g2:02X}{b2:02X}"


def _week_number_for_monday(monday: date, week_start_offset: int) -> int:
    all_mondays = get_mondays_in_month(monday.year, monday.month)
    try:
        i = all_mondays.index(monday)
        return ((i + week_start_offset - 1) % 4) + 1
    except ValueError:
        return week_start_offset


def _build_planned_week_rows(monday: date, week_start_offset: int) -> list[dict]:
    monday_str = str(monday)
    friday = monday + timedelta(days=4)
    wn = _week_number_for_monday(monday, week_start_offset)

    people_res = supabase.table("people").select("id, name").eq("active", True).order("name").execute()
    all_people = people_res.data

    abs_rows = supabase.table("absences").select("person_id, date, reported_by").gte(
        "date", monday_str
    ).lte("date", str(friday)).execute().data
    holiday_dates = holiday_dates_from_absence_rows(abs_rows)
    holiday_dows = holiday_dows_for_week(monday, holiday_dates)

    abs_by_pid: dict[str, set[str]] = defaultdict(set)
    for row in abs_rows:
        abs_by_pid[row["person_id"]].add(row["date"])

    bulk_sched_raw = supabase.table("person_schedule").select(
        "person_id, day_of_week, hours, location, valid_from, valid_until"
    ).execute().data
    sched_rows_by_pid: dict[str, list[dict]] = defaultdict(list)
    for row in bulk_sched_raw:
        sched_rows_by_pid[row["person_id"]].append(row)

    active_sched_rows = []
    for person_rows in sched_rows_by_pid.values():
        active_sched_rows.extend(active_schedule_rows(person_rows, monday_str))

    bulk_dist_raw = supabase.table("task_distribution").select(
        "person_id, task_id, hours_per_week, preferred_days, valid_from, "
        "tasks(id, name, color, responsible_person, schedule_rule, is_fill, priority)"
    ).eq("week_number", wn).execute().data
    bulk_dist = active_distribution_rows(bulk_dist_raw, monday_str)

    bulk_realloc = supabase.table("temporary_reallocations").select(
        "week_start_date, covering_person_id, task_id, hours, "
        "task:task_id(id, name, color, responsible_person, schedule_rule, is_fill, priority)"
    ).eq("week_start_date", monday_str).execute().data

    tp_res = supabase.table("task_people").select(
        "person_id, task_id, day_hours"
    ).eq("week_number", wn).execute()
    day_hours_by_person_task = {
        (row["person_id"], row["task_id"]): row["day_hours"]
        for row in tp_res.data
        if row.get("day_hours")
    }

    sched_by_pid: dict[str, list[dict]] = defaultdict(list)
    for row in active_sched_rows:
        sched_by_pid[row["person_id"]].append(row)

    dist_by_pid: dict[str, list[dict]] = defaultdict(list)
    for row in bulk_dist:
        dist_by_pid[row["person_id"]].append(row)

    realloc_by_pid: dict[str, list[dict]] = defaultdict(list)
    for row in bulk_realloc:
        realloc_by_pid[row["covering_person_id"]].append(row)

    rows_to_insert = []
    for person in all_people:
        pid = person["id"]
        pname = person["name"]

        week_schedule = {
            row["day_of_week"]: row["hours"]
            for row in sched_by_pid[pid]
            if row["hours"] > 0 and row["day_of_week"] not in holiday_dows
        }
        if not week_schedule:
            continue

        tasks_list = []
        preferred_days = {}
        for row in dist_by_pid[pid]:
            tasks_list.append({
                "task_id": row["task_id"],
                "task_name": row["tasks"]["name"],
                "task_color": row["tasks"].get("color"),
                "responsible_person": row["tasks"].get("responsible_person"),
                "hours_per_week": row["hours_per_week"],
                "schedule_rule": row["tasks"].get("schedule_rule"),
                "is_fill": row["tasks"].get("is_fill", False),
                "priority": row["tasks"].get("priority"),
            })
            if row.get("preferred_days"):
                preferred_days[row["task_id"]] = row["preferred_days"]

        tasks_list = _overlay_reallocations_on_tasks(tasks_list, realloc_by_pid[pid])
        task_map = {task["task_id"]: task for task in tasks_list}

        person_day_hours = {}
        for row in dist_by_pid[pid]:
            day_hours = day_hours_by_person_task.get((pid, row["task_id"]))
            if day_hours:
                person_day_hours[row["task_id"]] = {int(k): v for k, v in day_hours.items()}

        allocations, _warnings = distribute_week(
            tasks_list,
            week_schedule,
            pname,
            preferred_days,
            day_hours_map=person_day_hours or None,
            blocked_days=holiday_dows,
        )

        for dow in range(1, 6):
            day_date = monday + timedelta(days=dow - 1)
            day_str = str(day_date)
            if day_str in abs_by_pid[pid]:
                continue
            if not week_schedule.get(dow, 0):
                continue

            for task_id, hours in allocations.get(dow, {}).items():
                if hours <= 0:
                    continue
                task_meta = task_map.get(task_id)
                if not task_meta:
                    continue
                rows_to_insert.append({
                    "person_id": pid,
                    "task_id": task_id,
                    "task_label": task_meta["task_name"],
                    "date": day_str,
                    "hours": hours,
                })

    return rows_to_insert


@router.get("")
def get_actual(week_start: str = Query(...)):
    """Return all actual_hours rows for Mon–Fri of the given week."""
    monday = date.fromisoformat(week_start)
    friday = monday + timedelta(days=4)
    res = (
        supabase.table("actual_hours")
        .select("id, person_id, task_id, task_label, date, hours, created_at, people(id, name)")
        .gte("date", str(monday))
        .lte("date", str(friday))
        .order("date")
        .execute()
    )
    return res.data


@router.post("", status_code=201)
def create_actual(body: ActualHoursCreate):
    row = {
        "person_id": body.person_id,
        "task_label": body.task_label,
        "date": str(body.date),
        "hours": body.hours,
    }
    if body.task_id:
        row["task_id"] = body.task_id
    res = supabase.table("actual_hours").insert(row).execute()
    return res.data[0]


@router.put("/{entry_id}")
def update_actual(entry_id: str, body: ActualHoursUpdate):
    patch = {}
    if body.hours is not None:
        patch["hours"] = body.hours
    if body.task_label is not None:
        patch["task_label"] = body.task_label
    if body.date is not None:
        patch["date"] = str(body.date)
    res = supabase.table("actual_hours").update(patch).eq("id", entry_id).execute()
    return res.data[0]


@router.delete("/{entry_id}", status_code=204)
def delete_actual(entry_id: str):
    supabase.table("actual_hours").delete().eq("id", entry_id).execute()
    return Response(status_code=204)


@router.get("/location")
def get_actual_location(week_start: str = Query(...)):
    """Return actual_location rows for Mon–Fri of the given week, merged with schedule defaults."""
    monday = date.fromisoformat(week_start)
    friday = monday + timedelta(days=4)

    # Explicit overrides for this week
    overrides = (
        supabase.table("actual_location")
        .select("person_id, date, location")
        .gte("date", str(monday))
        .lte("date", str(friday))
        .execute()
        .data
    )
    override_map = {(r["person_id"], r["date"]): r["location"] for r in overrides}

    # Schedule defaults (location per person per day_of_week)
    sched_rows = (
        supabase.table("person_schedule")
        .select("person_id, day_of_week, location, valid_from, valid_until")
        .execute()
        .data
    )
    _sched_by_pid: dict = defaultdict(list)
    for r in sched_rows:
        _sched_by_pid[r["person_id"]].append(r)
    sched_rows = []
    for _pid_rows in _sched_by_pid.values():
        sched_rows.extend(active_schedule_rows(_pid_rows, str(monday)))
    sched_map = {}  # person_id -> {day_of_week: location}
    for r in sched_rows:
        sched_map.setdefault(r["person_id"], {})[r["day_of_week"]] = r.get("location") or "office"

    # Build result: for every person × workday, return effective location.
    # Non-working days should stay empty instead of defaulting to office.
    people = supabase.table("people").select("id").eq("active", True).execute().data
    result = {}
    for p in people:
        pid = p["id"]
        result[pid] = {}
        for i in range(5):
            d = monday + timedelta(days=i)
            d_str = str(d)
            dow = i + 1  # 1=Mon…5=Fri
            default = sched_map.get(pid, {}).get(dow)
            if default is None:
                result[pid][d_str] = None
            else:
                result[pid][d_str] = override_map.get((pid, d_str), default)

    return result


@router.put("/location")
def upsert_actual_location(body: ActualLocationUpsert):
    supabase.table("actual_location").upsert(
        {"person_id": body.person_id, "date": str(body.date), "location": body.location},
        on_conflict="person_id,date",
    ).execute()
    return {"ok": True}


def _build_actual_excel_response(date_list: list[str]):
    if not date_list:
        return Response(status_code=400)

    rows = (
        supabase.table("actual_hours")
        .select("person_id, task_id, task_label, date, hours, people(name), tasks(name, color)")
        .in_("date", date_list)
        .execute()
        .data
    )

    # Collect all person names (sorted alphabetically)
    person_names = sorted({r["people"]["name"] for r in rows if r.get("people")})

    # Styles
    HDR_F  = _sf("3730A3")
    TOT_F  = _sf("F0FDF4")
    WHITE  = _sf("FFFFFF")
    T = Side(style="thin",   color="D1D5DB")
    NO = Side(style=None)
    thin_border = Border(left=T, right=T, top=T, bottom=T)
    no_left = Border(left=NO, right=T, top=T, bottom=T)

    DAY_ABBR = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for date_str in date_list:
        d_obj = date.fromisoformat(date_str)
        sheet_name = f"{DAY_ABBR[d_obj.weekday()]} {d_obj.day:02d} {MONTH_ABBR[d_obj.month - 1]}"
        ws = wb.create_sheet(title=sheet_name)

        day_rows = [r for r in rows if r["date"] == date_str]

        # Group: task_key → {task_label, task_color, person_name → hours}
        task_order = []
        task_data = {}
        for r in day_rows:
            pname = r["people"]["name"] if r.get("people") else "?"
            tname = (r["tasks"]["name"] if r.get("tasks") and r["tasks"] else None) or r["task_label"]
            tcolor = (r["tasks"].get("color") or "").lstrip("#") if r.get("tasks") and r["tasks"] else ""
            key = r["task_id"] or f"__adhoc__{r['task_label']}"
            if key not in task_data:
                task_order.append(key)
                task_data[key] = {"label": tname, "color": tcolor, "people": {}}
            task_data[key]["people"][pname] = task_data[key]["people"].get(pname, 0) + r["hours"]

        # Header row: Task | [persons...] | Total
        headers = ["Task"] + person_names + ["Total"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HDR_F
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 18

        # Task rows
        for row_idx, key in enumerate(task_order, start=2):
            td = task_data[key]
            tcolor = td["color"] or "6366F1"
            light = _lighten(tcolor, 0.88)
            row_vals = [td["label"]] + [td["people"].get(p, None) for p in person_names]
            row_vals.append(None)
            ws.append(row_vals)

            # Task name cell
            name_cell = ws.cell(row=row_idx, column=1)
            name_cell.fill = _sf(light)
            name_cell.font = Font(color=tcolor, size=9, bold=False)
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            name_cell.border = thin_border

            # Person + total cells
            for col_idx in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = _sf(light) if cell.value else WHITE
                cell.font = Font(size=9, bold=(col_idx == len(headers)))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            if person_names:
                total_cell = ws.cell(row=row_idx, column=len(headers))
                total_cell.value = f"=SUM(B{row_idx}:{get_column_letter(len(headers) - 1)}{row_idx})"
            ws.row_dimensions[row_idx].height = 15

        # Team Total row
        total_row_idx = len(task_order) + 2
        totals = ["Team Total"] + ([None] * len(person_names)) + [None]
        ws.append(totals)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = TOT_F
            cell.font = Font(bold=True, color="166534", size=9)
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            cell.border = thin_border
        for col_idx in range(2, len(headers) + 1):
            ws.cell(row=total_row_idx, column=col_idx).value = f"=SUM({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{total_row_idx - 1})"
        ws.row_dimensions[total_row_idx].height = 16

        # Column widths
        ws.column_dimensions["A"].width = 26
        for ci in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 10

        ws.freeze_panes = "B2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    first, last = date_list[0], date_list[-1]
    filename = f"actual_{first}_to_{last}.xlsx" if first != last else f"actual_{first}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export")
def export_actual_excel(dates: str = Query(...)):
    """Download an Excel file with one sheet per date showing actual hours (Task × Person)."""
    date_list = sorted(set(d.strip() for d in dates.split(",") if d.strip()))
    return _build_actual_excel_response(date_list)


@router.get("/export-month")
def export_actual_month_excel(year: int = Query(...), month: int = Query(..., ge=1, le=12)):
    """Download an Excel file with one sheet per workday for a whole month."""
    last_day = cal_module.monthrange(year, month)[1]
    date_list = [
        str(date(year, month, day))
        for day in range(1, last_day + 1)
        if date(year, month, day).weekday() < 5
    ]
    return _build_actual_excel_response(date_list)


@router.post("/copy-week")
def copy_week(body: CopyWeekRequest):
    """Populate actual_hours from the planner's weekly allocations for the selected week."""
    monday = body.week_start
    friday = monday + timedelta(days=4)

    # Check if data already exists for this week
    existing = (
        supabase.table("actual_hours")
        .select("id", count="exact")
        .gte("date", str(monday))
        .lte("date", str(friday))
        .execute()
    )
    if (existing.count or 0) > 0:
        if not body.force:
            return {"created": 0, "skipped": True}
        # force=True: delete existing rows and re-copy fresh
        supabase.table("actual_hours").delete().gte("date", str(monday)).lte("date", str(friday)).execute()

    rows_to_insert = _build_planned_week_rows(monday, body.week_start_offset)

    if rows_to_insert:
        supabase.table("actual_hours").insert(rows_to_insert).execute()

    return {"created": len(rows_to_insert), "skipped": False}
