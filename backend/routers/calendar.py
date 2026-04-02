from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from database import supabase
from datetime import date, timedelta
from io import BytesIO
import calendar as cal_module
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.versioned import active_schedule_rows, active_distribution_rows

router = APIRouter(prefix="/calendar", tags=["calendar"])

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri"]


def round_half(x: float) -> float:
    return round(x * 2) / 2


def norm(name: str) -> str:
    return name.lower().strip()


# ── Rule-based distribution ───────────────────────────────────────────────────

RULE_PRIORITY = {
    "do_not_split":  1,
    "one_day":       2,
    "first_work_day": 3,
    "two_days":      4,
    "flexible_days": 5,
    "proportional":  6,
    "equal_per_day": 7,
    None:            8,
}


def get_rule_priority(task: dict) -> tuple:
    rule = task.get("schedule_rule")
    return (RULE_PRIORITY.get(rule, 8), task.get("priority") or 99)


# ── Distribution engine ───────────────────────────────────────────────────────

def distribute_week(
    tasks: list[dict],       # [{task_id, task_name, schedule_rule, hours_per_week, is_fill, priority}]
    schedule: dict,          # {dow (1–5): hours}  — only non-zero days
    person_name: str,
    preferred_days: dict = None,  # {task_id: list[int]} preferred dows — overrides rules
) -> tuple[dict, list[str]]:     # ({dow: {task_id: hours}}, warnings)
    """
    Distribute one person's weekly task hours across their work days using
    the task's schedule_rule. Returns (allocations, warnings).
    """
    work_dows = sorted(schedule.keys())
    warnings: list[str] = []
    if not work_dows or not tasks:
        return {dow: {} for dow in work_dows}, warnings

    day_capacity = {dow: schedule[dow] for dow in work_dows}
    allocations: dict[int, dict[str, float]] = {dow: {} for dow in work_dows}
    total_sched = sum(schedule[d] for d in work_dows)

    def top_by_capacity(n: int) -> list[int]:
        return sorted(work_dows, key=lambda d: day_capacity[d], reverse=True)[:n]

    def alloc(dow: int, task_id: str, hours: float):
        hours = round_half(hours)
        if hours <= 0:
            return
        allocations[dow][task_id] = allocations[dow].get(task_id, 0.0) + hours
        day_capacity[dow] = max(0.0, day_capacity[dow] - hours)

    normal_tasks = sorted([t for t in tasks if not t.get("is_fill")], key=get_rule_priority)
    fill_tasks   = [t for t in tasks if t.get("is_fill")]

    for t in normal_tasks:
        tid = t["task_id"]
        hrs = t["hours_per_week"]
        if hrs <= 0:
            continue

        # ── Preferred day pin: overrides all rules ──
        if preferred_days and tid in preferred_days:
            pinned = preferred_days[tid]
            if isinstance(pinned, int):  # backward-compat
                pinned = [pinned]
            valid = [d for d in pinned if d in work_dows]
            if valid:
                rem = hrs
                for i, dow in enumerate(valid):
                    if i == len(valid) - 1:
                        alloc(dow, tid, round_half(rem))
                    else:
                        share = round_half(hrs / len(valid))
                        alloc(dow, tid, share)
                        rem = round_half(rem - share)
            else:
                best = top_by_capacity(1)
                if best:
                    alloc(best[0], tid, hrs)
            continue

        rule = t.get("schedule_rule")

        # ── Do not split: all hours on one day, warn if capacity insufficient ──
        if rule == "do_not_split":
            best = top_by_capacity(1)
            if best and day_capacity[best[0]] >= hrs:
                alloc(best[0], tid, hrs)
            else:
                warnings.append(f"{t['task_name']} ({person_name}): could not keep on one day — not enough capacity")
                if best:
                    alloc(best[0], tid, hrs)

        # ── One day: all hours on best available day ──
        elif rule == "one_day":
            best = top_by_capacity(1)
            if best:
                alloc(best[0], tid, hrs)

        # ── First work day: all hours on first working day of person's week ──
        elif rule == "first_work_day":
            alloc(work_dows[0], tid, hrs)

        # ── Two days: split across 2 best-capacity days ──
        elif rule == "two_days":
            top2 = top_by_capacity(2)
            if len(top2) >= 2:
                half = round_half(hrs / 2)
                alloc(top2[0], tid, half)
                alloc(top2[1], tid, hrs - half)
            elif top2:
                alloc(top2[0], tid, hrs)

        # ── Flexible days: 2 days if capacity fits, expand to 3 if needed ──
        elif rule == "flexible_days":
            top2 = top_by_capacity(2)
            if len(top2) >= 2 and sum(day_capacity[d] for d in top2) >= hrs:
                half = round_half(hrs / 2)
                alloc(top2[0], tid, half)
                alloc(top2[1], tid, hrs - half)
            else:
                top3 = top_by_capacity(min(3, len(work_dows)))
                rem = hrs
                for i, dow in enumerate(top3):
                    day_hrs = round_half(hrs / len(top3)) if i < len(top3) - 1 else round_half(rem)
                    alloc(dow, tid, day_hrs)
                    rem = round_half(rem - day_hrs)

        # ── Equal per day: same hours every work day ──
        elif rule == "equal_per_day":
            rem = hrs
            for i, dow in enumerate(work_dows):
                day_hrs = round_half(hrs / len(work_dows)) if i < len(work_dows) - 1 else round_half(rem)
                alloc(dow, tid, day_hrs)
                rem = round_half(rem - day_hrs)

        # ── Proportional or no rule (default): proportional to scheduled hours ──
        else:
            rem = hrs
            for i, dow in enumerate(work_dows):
                if i == len(work_dows) - 1:
                    day_hrs = round_half(rem)
                else:
                    day_hrs = round_half(hrs * schedule[dow] / total_sched)
                    rem = round_half(rem - day_hrs)
                alloc(dow, tid, day_hrs)

    # ── Fill tasks: absorb remaining daily capacity ──
    for t in fill_tasks:
        tid = t["task_id"]
        for dow in work_dows:
            cap = round_half(day_capacity[dow])
            if cap > 0:
                alloc(dow, tid, cap)

    return allocations, warnings


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_mondays_in_month(year: int, month: int) -> list[date]:
    first_day = date(year, month, 1)
    last_day  = date(year, month, cal_module.monthrange(year, month)[1])
    # First Monday that actually falls within this month
    first_monday = first_day + timedelta(days=(7 - first_day.weekday()) % 7)
    mondays = []
    current = first_monday
    while current <= last_day:
        mondays.append(current)
        current += timedelta(weeks=1)
    return mondays


# ── Endpoint ──────────────────────────────────────────────────────────────────

def distribute_week_proportional(
    tasks: list[dict],
    schedule: dict,
) -> dict:
    """Simple proportional distribution (no rules) — used for locked weeks."""
    work_dows = sorted(schedule.keys())
    if not work_dows or not tasks:
        return {dow: {} for dow in work_dows}
    total_sched = sum(schedule[d] for d in work_dows)
    allocations: dict[int, dict[str, float]] = {dow: {} for dow in work_dows}
    for t in tasks:
        tid = t["task_id"]
        hrs = t["hours_per_week"]
        if hrs <= 0:
            continue
        rem = hrs
        for i, dow in enumerate(work_dows):
            if i == len(work_dows) - 1:
                day_hrs = round_half(rem)
            else:
                day_hrs = round_half(hrs * schedule[dow] / total_sched)
                rem = round_half(rem - day_hrs)
            if day_hrs > 0:
                allocations[dow][tid] = day_hrs
    return allocations


MONTH_NAMES_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_NAMES_LONG  = ["January","February","March","April","May","June",
                      "July","August","September","October","November","December"]
DAY_ABBR = ["Mon","Tue","Wed","Thu","Fri"]


def _compute_day_view(date_obj: date, week_start: int = 1) -> dict:
    """Shared logic for GET /day and GET /day/export."""
    from collections import defaultdict

    dow = date_obj.weekday() + 1  # 1=Mon … 5=Fri
    monday = date_obj - timedelta(days=date_obj.weekday())
    year, month = date_obj.year, date_obj.month

    # Determine week_number (1–4) relative to the month
    all_mondays = get_mondays_in_month(year, month)
    try:
        i = all_mondays.index(monday)
        wn = ((i + week_start - 1) % 4) + 1
    except ValueError:
        wn = week_start

    # ── Bulk queries (4 total) ────────────────────────────────────────────────
    people_res = supabase.table("people").select("id, name").eq("active", True).order("name").execute()
    all_people = people_res.data

    abs_res = supabase.table("absences").select("person_id").eq("date", str(date_obj)).execute()
    absent_ids = {r["person_id"] for r in abs_res.data}

    monday_str = str(monday)
    bulk_sched_raw = supabase.table("person_schedule").select("person_id, day_of_week, hours, location, valid_from, valid_until").execute().data
    bulk_sched = active_schedule_rows(bulk_sched_raw, monday_str)
    bulk_dist_raw  = supabase.table("task_distribution").select(
        "person_id, task_id, hours_per_week, preferred_days, valid_from, tasks(id, name, color, responsible_person, schedule_rule, is_fill, priority)"
    ).eq("week_number", wn).execute().data
    bulk_dist = active_distribution_rows(bulk_dist_raw, monday_str)

    sched_by_pid: dict = defaultdict(list)
    for r in bulk_sched:
        sched_by_pid[r["person_id"]].append(r)

    dist_by_pid: dict = defaultdict(list)
    for r in bulk_dist:
        dist_by_pid[r["person_id"]].append(r)

    # Per-person allocations for the day
    # task_id -> {task_name, task_color, responsible_person, people: [{person_name, hours}]}
    task_data: dict[str, dict] = {}

    for person in all_people:
        pid   = person["id"]
        pname = person["name"]

        schedule = {r["day_of_week"]: r["hours"] for r in sched_by_pid[pid] if r["hours"] > 0}
        location_map = {r["day_of_week"]: (r.get("location") or "office") for r in sched_by_pid[pid]}

        if not schedule.get(dow, 0) or pid in absent_ids:
            continue  # person doesn't work this day or is absent

        location = location_map.get(dow, "office")

        tasks_list = []
        preferred  = {}
        for row in dist_by_pid[pid]:
            tasks_list.append({
                "task_id":        row["task_id"],
                "task_name":      row["tasks"]["name"],
                "task_color":     row["tasks"].get("color"),
                "responsible_person": row["tasks"].get("responsible_person"),
                "hours_per_week": row["hours_per_week"],
                "schedule_rule":  row["tasks"].get("schedule_rule"),
                "is_fill":        row["tasks"].get("is_fill", False),
                "priority":       row["tasks"].get("priority"),
            })
            if row.get("preferred_days"):
                preferred[row["task_id"]] = row["preferred_days"]

        alloc, _warnings = distribute_week(tasks_list, schedule, pname, preferred)

        for tid, hrs in alloc.get(dow, {}).items():
            if hrs <= 0:
                continue
            task_meta = next((t for t in tasks_list if t["task_id"] == tid), None)
            if not task_meta:
                continue
            if tid not in task_data:
                task_data[tid] = {
                    "task_id":            tid,
                    "task_name":          task_meta["task_name"],
                    "task_color":         task_meta["task_color"],
                    "responsible_person": task_meta["responsible_person"],
                    "total_hours":        0.0,
                    "people":             [],
                }
            task_data[tid]["total_hours"] = round(task_data[tid]["total_hours"] + hrs, 2)
            task_data[tid]["people"].append({"person_name": pname, "hours": hrs, "location": location})

    tasks_sorted = sorted(task_data.values(), key=lambda t: t["total_hours"], reverse=True)
    absent_names = [p["name"] for p in all_people if p["id"] in absent_ids]
    total_hours  = round(sum(t["total_hours"] for t in tasks_sorted), 2)

    return {
        "date":         str(date_obj),
        "day_name":     DAY_ABBR[date_obj.weekday()],
        "week_number":  wn,
        "tasks":        tasks_sorted,
        "absent_people": absent_names,
        "total_hours":  total_hours,
    }


@router.get("/day")
def get_day_view(date_str: str = Query(..., alias="date"), week_start: int = Query(default=1, ge=1, le=4)):
    """Return all tasks and assigned people with hours for a single day."""
    date_obj = date.fromisoformat(date_str)
    if date_obj.weekday() >= 5:
        return {"date": date_str, "day_name": DAY_ABBR[min(date_obj.weekday(), 4)],
                "week_number": None, "tasks": [], "absent_people": [], "total_hours": 0,
                "is_weekend": True}
    return _compute_day_view(date_obj, week_start)


@router.get("/day/export")
def export_day_excel(date_str: str = Query(..., alias="date"), week_start: int = Query(default=1, ge=1, le=4)):
    """Download an Excel file for a single day — tasks × people."""
    date_obj = date.fromisoformat(date_str)
    view = _compute_day_view(date_obj, week_start)

    # Collect all person names (column headers)
    person_names: list[str] = []
    seen: set[str] = set()
    for t in view["tasks"]:
        for p in t["people"]:
            if p["person_name"] not in seen:
                person_names.append(p["person_name"])
                seen.add(p["person_name"])

    wb = openpyxl.Workbook()
    ws = wb.active
    day_label = f"{view['day_name']} {date_obj.day} {MONTH_NAMES_SHORT[date_obj.month-1]} {date_obj.year}"
    ws.title = day_label[:31]

    HDR_FILL  = PatternFill("solid", fgColor="3730A3")
    BOLD_WHITE = Font(bold=True, color="FFFFFF")
    BOLD      = Font(bold=True)
    centre    = Alignment(horizontal="center")

    # Header
    header = ["Task", "Responsible"] + person_names + ["Total"]
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = BOLD_WHITE
        cell.fill = HDR_FILL
        cell.alignment = centre

    # Task rows
    for t in view["tasks"]:
        person_hours = {p["person_name"]: p["hours"] for p in t["people"]}
        row = [t["task_name"], t["responsible_person"] or "—"]
        row += [person_hours.get(pname) for pname in person_names]
        row += [t["total_hours"]]
        ws.append(row)
        # Color the task name cell
        if t["task_color"]:
            color = t["task_color"].lstrip("#")
            ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=ws.max_row, column=len(header)).font = BOLD

    # Absent footer
    if view["absent_people"]:
        ws.append([])
        ws.append([f"Absent: {', '.join(view['absent_people'])}"])
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="991B1B")

    # Widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 15
    for col in range(3, len(header) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.freeze_panes = "C2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"daily_{date_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-data")
def get_calendar_export_data(year: int = Query(...), month: int = Query(...), week_start: int = Query(default=1, ge=1, le=4)):
    """
    Return pre-computed daily allocations for all active people as JSON.
    Used by the browser to generate the Excel file client-side.
    """
    from collections import defaultdict

    prev_m = month - 1 if month > 1 else 12
    prev_y = year if month > 1 else year - 1
    prev_last      = date(prev_y, prev_m, cal_module.monthrange(prev_y, prev_m)[1])
    last_monday_prev = prev_last - timedelta(days=prev_last.weekday())

    cur_first = date(year, month, 1)
    cur_last  = date(year, month, cal_module.monthrange(year, month)[1])

    section_a = [last_monday_prev + timedelta(days=i) for i in range(5)]
    section_b: list[date] = []
    all_mondays = get_mondays_in_month(year, month)
    for monday in all_mondays:
        if monday <= last_monday_prev:
            continue
        for dow in range(1, 6):
            d = monday + timedelta(days=dow - 1)
            if d.month == month:
                section_b.append(d)

    # ── Bulk queries ────────────────────────────────────────────────────────
    people_res = supabase.table("people").select("id, name").eq("active", True).order("name").execute()
    all_people = people_res.data

    bulk_sched_raw = supabase.table("person_schedule").select("person_id, day_of_week, hours, location, valid_from, valid_until").execute().data
    bulk_dist_raw  = supabase.table("task_distribution").select(
        "person_id, week_number, task_id, hours_per_week, preferred_days, valid_from, tasks(id, name, color, schedule_rule, is_fill, priority)"
    ).execute().data
    bulk_abs   = supabase.table("absences").select("person_id, date").gte(
        "date", str(last_monday_prev)
    ).lte("date", str(cur_last)).execute().data

    sched_by_pid: dict = defaultdict(list)
    for r in bulk_sched_raw:
        sched_by_pid[r["person_id"]].append(r)

    dist_by_pid: dict = defaultdict(list)
    for r in bulk_dist_raw:
        dist_by_pid[r["person_id"]].append(r)
        tc = (r["tasks"].get("color") or "").lstrip("#")

    abs_by_pid: dict = defaultdict(set)
    for r in bulk_abs:
        abs_by_pid[r["person_id"]].add(r["date"])

    # ── Per-person allocations ───────────────────────────────────────────────
    person_alloc:   dict = {}
    task_color_map: dict = {}

    def _tasks_for(pid, wn, monday_str):
        rows = active_distribution_rows(
            [r for r in dist_by_pid[pid] if r["week_number"] == wn], monday_str
        )
        for r in rows:
            tc = (r["tasks"].get("color") or "").lstrip("#")
            if tc and r["tasks"]["name"] not in task_color_map:
                task_color_map[r["tasks"]["name"]] = tc
        tasks = [{"task_id": r["task_id"], "task_name": r["tasks"]["name"],
                  "task_color": r["tasks"].get("color"), "hours_per_week": r["hours_per_week"],
                  "schedule_rule": r["tasks"].get("schedule_rule"),
                  "is_fill": r["tasks"].get("is_fill", False),
                  "priority": r["tasks"].get("priority")} for r in rows]
        preferred = {r["task_id"]: r["preferred_days"] for r in rows if r.get("preferred_days")}
        return tasks, preferred

    def _sched_for(pid, monday_str):
        rows = active_schedule_rows(sched_by_pid[pid], monday_str)
        return {r["day_of_week"]: r["hours"] for r in rows if r["hours"] > 0}

    for person in all_people:
        pid   = person["id"]
        pname = person["name"]
        absent = abs_by_pid[pid]
        day_alloc: dict = {}

        if section_a:
            mon_a_str = str(last_monday_prev)
            tasks_a, preferred_a = _tasks_for(pid, 4, mon_a_str)
            schedule_a = _sched_for(pid, mon_a_str)
            tmap_a = {t["task_id"]: t for t in tasks_a}
            alloc_a, _warnings = distribute_week(tasks_a, schedule_a, pname, preferred_a)
            for d in section_a:
                d_str = str(d)
                dow   = d.weekday() + 1
                if d_str in absent:
                    day_alloc[d_str] = "absent"
                elif schedule_a.get(dow, 0) > 0:
                    day_alloc[d_str] = {tmap_a[tid]["task_name"]: hrs for tid, hrs in alloc_a.get(dow, {}).items() if hrs > 0 and tid in tmap_a}
                else:
                    day_alloc[d_str] = None

        for i, monday in enumerate(all_mondays):
            if monday <= last_monday_prev:
                continue
            wn         = ((i + week_start - 1) % 4) + 1
            mon_str    = str(monday)
            tasks_w, preferred_w = _tasks_for(pid, wn, mon_str)
            schedule_w = _sched_for(pid, mon_str)
            tmap_w     = {t["task_id"]: t for t in tasks_w}
            week_sched = {dow: hrs for dow, hrs in schedule_w.items() if (monday + timedelta(days=dow - 1)).month == month}
            alloc_w, _warnings = distribute_week(tasks_w, week_sched, pname, preferred_w)
            for dow in range(1, 6):
                d = monday + timedelta(days=dow - 1)
                if d.month != month:
                    continue
                d_str = str(d)
                if d_str in absent:
                    day_alloc[d_str] = "absent"
                elif schedule_w.get(dow, 0) > 0:
                    day_alloc[d_str] = {tmap_w[tid]["task_name"]: hrs for tid, hrs in alloc_w.get(dow, {}).items() if hrs > 0 and tid in tmap_w}
                else:
                    day_alloc[d_str] = None

        person_alloc[pid] = day_alloc

    # ── Week group metadata (for frontend column colouring) ──────────────────
    WK_COLORS = ["3730A3", "1E40AF", "0369A1", "0F766E"]
    groups_data: list = []
    if section_a:
        groups_data.append({"label": "Prev Month", "start_idx": 0, "end_idx": len(section_a) - 1, "color": "6D28D9"})
    prev_mon = None
    wk_idx   = 0
    wk_start_idx = len(section_a)
    for col_off, d in enumerate(section_b):
        mon = d - timedelta(days=d.weekday())
        if mon != prev_mon:
            if prev_mon is not None:
                groups_data.append({"label": f"Week {wk_idx}", "start_idx": wk_start_idx, "end_idx": len(section_a) + col_off - 1, "color": WK_COLORS[(wk_idx - 1) % 4]})
            wk_start_idx = len(section_a) + col_off
            prev_mon = mon
            wk_idx  += 1
    if section_b:
        groups_data.append({"label": f"Week {wk_idx}", "start_idx": wk_start_idx, "end_idx": len(section_a) + len(section_b) - 1, "color": WK_COLORS[(wk_idx - 1) % 4]})

    return {
        "year": year,
        "month": month,
        "month_name": MONTH_NAMES_LONG[month - 1],
        "cur_first": str(cur_first),
        "section_a": [str(d) for d in section_a],
        "section_b": [str(d) for d in section_b],
        "groups": groups_data,
        "people": [{"id": p["id"], "name": p["name"]} for p in all_people],
        "allocations": {pid: {k: v for k, v in alloc.items()} for pid, alloc in person_alloc.items()},
        "task_colors": task_color_map,
    }


@router.get("/export")
def export_calendar_excel(year: int = Query(...), month: int = Query(...), week_start: int = Query(default=1, ge=1, le=4)):
    """
    Download an Excel file with all active people's daily task hours for the
    selected month, including the full last working week of the previous month
    as leading columns.
    """
    # ── Previous month ───────────────────────────────────────────────────────
    prev_m = month - 1 if month > 1 else 12
    prev_y = year if month > 1 else year - 1
    prev_last = date(prev_y, prev_m, cal_module.monthrange(prev_y, prev_m)[1])
    last_monday_prev = prev_last - timedelta(days=prev_last.weekday())

    cur_first = date(year, month, 1)
    cur_last  = date(year, month, cal_module.monthrange(year, month)[1])

    # ── Section A: full last Mon–Fri week of previous month ─────────────────
    section_a = [last_monday_prev + timedelta(days=i) for i in range(5)]

    # ── Section B: working days of current month not in section A ────────────
    section_b = []
    all_mondays = get_mondays_in_month(year, month)
    for monday in all_mondays:
        if monday <= last_monday_prev:
            continue
        for dow in range(1, 6):
            d = monday + timedelta(days=dow - 1)
            if d.month == month:
                section_b.append(d)

    all_days = section_a + section_b

    # ── Bulk queries (4 total instead of 3×N) ────────────────────────────────
    people_res = supabase.table("people").select("id, name").eq("active", True).order("name").execute()
    all_people = people_res.data

    bulk_sched_raw2 = supabase.table("person_schedule").select("person_id, day_of_week, hours, location, valid_from, valid_until").execute().data
    bulk_dist_raw2  = supabase.table("task_distribution").select(
        "person_id, week_number, task_id, hours_per_week, preferred_days, valid_from, tasks(id, name, color, schedule_rule, is_fill, priority)"
    ).execute().data
    bulk_abs   = supabase.table("absences").select("person_id, date").gte(
        "date", str(last_monday_prev)
    ).lte("date", str(cur_last)).execute().data

    from collections import defaultdict
    sched_by_pid2 = defaultdict(list)
    for r in bulk_sched_raw2:
        sched_by_pid2[r["person_id"]].append(r)

    dist_by_pid2 = defaultdict(list)
    for r in bulk_dist_raw2:
        dist_by_pid2[r["person_id"]].append(r)

    abs_by_pid = defaultdict(set)
    for r in bulk_abs:
        abs_by_pid[r["person_id"]].add(r["date"])

    # ── Per-person allocations ────────────────────────────────────────────────
    person_alloc:   dict[str, dict] = {}
    task_color_map: dict[str, str]  = {}

    def _tasks_for2(pid, wn, monday_str):
        rows = active_distribution_rows(
            [r for r in dist_by_pid2[pid] if r["week_number"] == wn], monday_str
        )
        for r in rows:
            tname = r["tasks"]["name"]
            tc    = (r["tasks"].get("color") or "").lstrip("#")
            if tc and tname not in task_color_map:
                task_color_map[tname] = tc
        tasks = [{"task_id": r["task_id"], "task_name": r["tasks"]["name"],
                  "task_color": r["tasks"].get("color"), "hours_per_week": r["hours_per_week"],
                  "schedule_rule": r["tasks"].get("schedule_rule"),
                  "is_fill": r["tasks"].get("is_fill", False),
                  "priority": r["tasks"].get("priority")} for r in rows]
        preferred = {r["task_id"]: r["preferred_days"] for r in rows if r.get("preferred_days")}
        return tasks, preferred

    def _sched_for2(pid, monday_str):
        rows = active_schedule_rows(sched_by_pid2[pid], monday_str)
        return {r["day_of_week"]: r["hours"] for r in rows if r["hours"] > 0}

    for person in all_people:
        pid   = person["id"]
        pname = person["name"]
        absent = abs_by_pid[pid]
        day_alloc: dict[str, object] = {}

        if section_a:
            mon_a_str = str(last_monday_prev)
            tasks_a, preferred_a = _tasks_for2(pid, 4, mon_a_str)
            schedule_a = _sched_for2(pid, mon_a_str)
            tmap_a = {t["task_id"]: t for t in tasks_a}
            alloc_a, _warnings = distribute_week(tasks_a, schedule_a, pname, preferred_a)
            for d in section_a:
                d_str = str(d)
                dow   = d.weekday() + 1
                if d_str in absent:
                    day_alloc[d_str] = "absent"
                elif schedule_a.get(dow, 0) > 0:
                    day_alloc[d_str] = {
                        tmap_a[tid]["task_name"]: hrs
                        for tid, hrs in alloc_a.get(dow, {}).items()
                        if hrs > 0 and tid in tmap_a
                    }
                else:
                    day_alloc[d_str] = None

        for i, monday in enumerate(all_mondays):
            if monday <= last_monday_prev:
                continue
            wn         = ((i + week_start - 1) % 4) + 1
            mon_str    = str(monday)
            tasks_w, preferred_w = _tasks_for2(pid, wn, mon_str)
            schedule_w = _sched_for2(pid, mon_str)
            tmap_w     = {t["task_id"]: t for t in tasks_w}
            week_sched = {
                dow: hrs
                for dow, hrs in schedule_w.items()
                if (monday + timedelta(days=dow - 1)).month == month
            }
            alloc_w, _warnings = distribute_week(tasks_w, week_sched, pname, preferred_w)

            for dow in range(1, 6):
                d = monday + timedelta(days=dow - 1)
                if d.month != month:
                    continue
                d_str = str(d)
                if d_str in absent:
                    day_alloc[d_str] = "absent"
                elif schedule_w.get(dow, 0) > 0:
                    day_alloc[d_str] = {
                        tmap_w[tid]["task_name"]: hrs
                        for tid, hrs in alloc_w.get(dow, {}).items()
                        if hrs > 0 and tid in tmap_w
                    }
                else:
                    day_alloc[d_str] = None

        person_alloc[pid] = day_alloc

    # task_color_map was populated during the person-processing loop above
    def lighten(hex6: str, factor: float = 0.82) -> str:
        """Blend hex6 toward white by factor."""
        try:
            r = int(hex6[0:2], 16); g = int(hex6[2:4], 16); b = int(hex6[4:6], 16)
            return f"{int(r+(255-r)*factor):02X}{int(g+(255-g)*factor):02X}{int(b+(255-b)*factor):02X}"
        except Exception:
            return "F3F4F6"

    # ── Border helpers ────────────────────────────────────────────────────────
    T  = Side(style="thin",   color="D1D5DB")
    M  = Side(style="medium", color="6B7280")
    TK = Side(style="thick",  color="312E81")
    NO = Side(style=None)

    def bdr(top=T, bot=T, lft=T, rgt=T):
        return Border(top=top, bottom=bot, left=lft, right=rgt)

    # ── Fill helpers ──────────────────────────────────────────────────────────
    def sf(c): return PatternFill("solid", fgColor=c)

    TITLE_F  = sf("312E81")
    PREV_H   = sf("6D28D9")
    WK_FILLS = [sf("3730A3"), sf("1E40AF"), sf("0369A1"), sf("0F766E")]
    DAY_H    = sf("374151")
    NAME_F   = sf("EEF2FF")
    ABS_F    = sf("FEE2E2")
    OFF_F    = sf("F9FAFB")
    TOT_F    = sf("F0FDF4")
    WHITE_F  = sf("FFFFFF")
    PREV_D   = sf("F5F3FF")

    ctr  = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left",   vertical="center")

    # ── Determine day-column groups ───────────────────────────────────────────
    # groups: list of (label, col_start, col_end, fill)   (1-based, offset by 2 for A/B)
    groups: list[tuple] = []
    if section_a:
        groups.append(("Prev Month", 3, 3 + len(section_a) - 1, PREV_H))
    prev_mon = None
    wk_idx   = 0
    wk_start_col = 3 + len(section_a)
    for col_off, d in enumerate(section_b):
        mon = d - timedelta(days=d.weekday())
        if mon != prev_mon:
            if prev_mon is not None:
                groups.append((f"Week {wk_idx}", wk_start_col, 3 + len(section_a) + col_off - 1, WK_FILLS[(wk_idx - 1) % 4]))
            wk_start_col = 3 + len(section_a) + col_off
            prev_mon = mon
            wk_idx  += 1
    if section_b:
        groups.append((f"Week {wk_idx}", wk_start_col, 2 + len(all_days), WK_FILLS[(wk_idx - 1) % 4]))

    ncols = 2 + len(all_days)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAMES_LONG[month-1]} {year}"[:31]

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.append([""] * ncols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1)
    c.value     = f"{MONTH_NAMES_LONG[month-1]} {year}"
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = TITLE_F
    c.alignment = ctr
    ws.row_dimensions[1].height = 24

    # ── Row 2: Week group headers ─────────────────────────────────────────────
    ws.append([""] * ncols)
    ws.cell(2, 1).value = "Person"; ws.cell(2, 1).font = Font(bold=True, color="FFFFFF"); ws.cell(2, 1).fill = DAY_H; ws.cell(2, 1).alignment = left
    ws.cell(2, 2).value = "Task";   ws.cell(2, 2).font = Font(bold=True, color="FFFFFF"); ws.cell(2, 2).fill = DAY_H; ws.cell(2, 2).alignment = left
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1)
    for label, c1, c2, fill in groups:
        if c1 <= c2:
            ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        cell = ws.cell(2, c1)
        cell.value     = label
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = fill
        cell.alignment = ctr
    ws.row_dimensions[2].height = 16

    # ── Row 3: Day headers ────────────────────────────────────────────────────
    day_hdr_row = ["", ""]
    for d in all_days:
        day_hdr_row.append(f"{DAY_ABBR[d.weekday()]}  {d.day}")
    ws.append(day_hdr_row)
    for col in range(1, ncols + 1):
        cell = ws.cell(3, col)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = DAY_H if col <= 2 else ([g[3] for g in groups if g[1] <= col <= g[2]] or [DAY_H])[0]
        cell.alignment = ctr
        cell.border    = bdr()
    ws.row_dimensions[3].height = 15

    # Thick left border at start of current-month columns
    first_cur_col = next((3 + i for i, d in enumerate(all_days) if d >= cur_first), None)
    if first_cur_col:
        for r in range(2, 4):
            cell = ws.cell(r, first_cur_col)
            cell.border = Border(top=cell.border.top, bottom=cell.border.bottom,
                                 left=TK, right=cell.border.right)

    # ── Data rows ─────────────────────────────────────────────────────────────
    ROW_START = 4
    row_num   = ROW_START

    for person in all_people:
        pid       = person["id"]
        pname     = person["name"]
        day_alloc = person_alloc[pid]

        # Collect task names in first-appearance order
        task_names: list[str] = []
        seen: set[str] = set()
        for d in all_days:
            td = day_alloc.get(str(d))
            if isinstance(td, dict):
                for tname in td:
                    if tname not in seen:
                        task_names.append(tname)
                        seen.add(tname)

        n_task_rows = max(len(task_names), 1)
        p_start_row = row_num

        if not task_names:
            # No tasks — single row
            ws.append(["", "—"] + [None] * len(all_days))
            for col in range(1, ncols + 1):
                ws.cell(row_num, col).border = bdr()
                ws.cell(row_num, col).fill   = WHITE_F
            ws.cell(row_num, 1).fill = NAME_F
            row_num += 1
        else:
            for t_idx, tname in enumerate(task_names):
                tc_hex = task_color_map.get(tname, "")
                task_light = sf(lighten(tc_hex, 0.80)) if tc_hex else sf("F9FAFB")
                task_vlt   = sf(lighten(tc_hex, 0.91)) if tc_hex else WHITE_F
                tc_font    = tc_hex if tc_hex else "374151"

                row_data = ["", tname]
                for d in all_days:
                    td = day_alloc.get(str(d))
                    if td == "absent":
                        row_data.append("ABS")
                    elif isinstance(td, dict):
                        row_data.append(td.get(tname) or None)
                    else:
                        row_data.append(None)
                ws.append(row_data)

                # Task name cell (col B)
                tc = ws.cell(row_num, 2)
                tc.font      = Font(bold=False, color=tc_font, size=9)
                tc.fill      = task_light
                tc.alignment = left
                tc.border    = bdr()

                # Person name col (col A) — filled in after merge
                ws.cell(row_num, 1).fill   = NAME_F
                ws.cell(row_num, 1).border = bdr()

                # Day cells
                for col_off, d in enumerate(all_days):
                    col  = 3 + col_off
                    cell = ws.cell(row_num, col)
                    td   = day_alloc.get(str(d))
                    val  = cell.value

                    if td == "absent":
                        cell.fill  = ABS_F
                        cell.font  = Font(color="991B1B", size=9, bold=True)
                        cell.value = "ABS" if t_idx == 0 else None
                    elif val:
                        cell.fill = task_vlt if d >= cur_first else sf(lighten(tc_hex or "A78BFA", 0.88))
                        cell.font = Font(color=tc_font, size=9)
                    elif schedule_entry := (day_alloc.get(str(d)) is None):
                        cell.fill = OFF_F
                    else:
                        cell.fill = OFF_F if d >= cur_first else PREV_D

                    cell.alignment = ctr
                    cell.border    = bdr(
                        lft=TK if d >= cur_first and col_off == (next((i for i,x in enumerate(all_days) if x >= cur_first), 0)) else T,
                        rgt=T
                    )

                ws.row_dimensions[row_num].height = 15
                row_num += 1

        # Merge person name vertically and write it
        if n_task_rows > 1:
            ws.merge_cells(start_row=p_start_row, start_column=1,
                           end_row=p_start_row + n_task_rows - 1, end_column=1)
        cell = ws.cell(p_start_row, 1)
        cell.value     = pname
        cell.font      = Font(bold=True, size=10, color="312E81")
        cell.fill      = NAME_F
        cell.alignment = Alignment(horizontal="left", vertical="top")

        # Thick bottom border after each person block
        for col in range(1, ncols + 1):
            c = ws.cell(row_num - 1, col)
            c.border = Border(top=c.border.top, bottom=M, left=c.border.left, right=c.border.right)

    # ── Totals row ────────────────────────────────────────────────────────────
    ws.append(["Total", ""] + [None] * len(all_days))
    tot_row = row_num
    for col_off, d in enumerate(all_days):
        d_str = str(d)
        total = 0.0
        for person in all_people:
            td = person_alloc[person["id"]].get(d_str)
            if isinstance(td, dict):
                total += sum(td.values())
        cell = ws.cell(tot_row, 3 + col_off)
        if total > 0:
            cell.value = total
        cell.fill      = TOT_F
        cell.font      = Font(bold=True, size=9, color="166534")
        cell.alignment = ctr
        cell.border    = bdr(top=M)
    ws.cell(tot_row, 1).font      = Font(bold=True, color="166534")
    ws.cell(tot_row, 1).fill      = TOT_F
    ws.cell(tot_row, 1).alignment = left
    ws.cell(tot_row, 1).border    = bdr(top=M)
    ws.cell(tot_row, 2).fill      = TOT_F
    ws.cell(tot_row, 2).border    = bdr(top=M)
    ws.row_dimensions[tot_row].height = 16

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    for col in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.freeze_panes = "C4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"calendar_{year}_{month:02d}_{MONTH_NAMES_LONG[month-1]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{year}/{month}")
def get_calendar(year: int, month: int, person_id: str = Query(...), from_week: int = Query(default=1, ge=1, le=5), week_start: int = Query(default=1, ge=1, le=4), include_overflow: bool = Query(default=False)):
    # Person info
    person_res = supabase.table("people").select("id, name").eq("id", person_id).single().execute()
    person = person_res.data

    # Schedule — all versioned rows for this person
    schedule_res = supabase.table("person_schedule").select("day_of_week, hours, valid_from, valid_until").eq("person_id", person_id).execute()
    all_sched_rows = [{**r, "person_id": person_id} for r in schedule_res.data]

    # Per-week distributions — all versioned rows for this person
    dist_res = supabase.table("task_distribution").select(
        "week_number, task_id, hours_per_week, preferred_days, valid_from, tasks(id, name, color, schedule_rule, is_fill, priority)"
    ).eq("person_id", person_id).execute()
    all_dist_rows = [{**r, "person_id": person_id} for r in dist_res.data]

    # We'll resolve per-week in the loop below; compute a baseline schedule for weekly_total
    _baseline_sched = active_schedule_rows(all_sched_rows, str(date(year, month, 1)))
    weekly_total = sum(r["hours"] for r in _baseline_sched)

    # Absences: extend range to cover overflow days in both directions
    first_day = date(year, month, 1)
    last_day  = date(year, month, cal_module.monthrange(year, month)[1])
    _month_mondays = get_mondays_in_month(year, month)
    # End: Friday of last week (may spill into next month, e.g. March 30 week → April 3)
    abs_end   = max(last_day, _month_mondays[-1] + timedelta(days=4)) if _month_mondays else last_day
    # Start: Monday of week before first Monday (covers include_overflow prev-month days)
    abs_start = (_month_mondays[0] - timedelta(weeks=1)) if _month_mondays else first_day
    absences_res = supabase.table("absences").select("date").eq("person_id", person_id).gte(
        "date", str(abs_start)
    ).lte("date", str(abs_end)).execute()
    absent_dates = {row["date"] for row in absences_res.data}

    # Build weeks
    mondays = get_mondays_in_month(year, month)

    # include_overflow: prepend prev month's last Monday as W1 (the "week 5 of prev month" becomes W1 here)
    if include_overflow and mondays:
        overflow_monday = mondays[0] - timedelta(weeks=1)
        # Only include if that Monday's week has workdays in this month
        if any((overflow_monday + timedelta(days=d)).month == month for d in range(1, 5)):
            mondays = [overflow_monday] + mondays

    weeks   = []

    for i, monday in enumerate(mondays):
        week_index  = i + 1
        # Overflow week (5th+ when not include_overflow) defaults to W2, not W1
        if i >= 4 and not include_overflow:
            week_number = 2
        else:
            week_number = ((i + week_start - 1) % 4) + 1

        monday_str = str(monday)
        # Resolve active schedule and distribution for this specific week
        active_sched = active_schedule_rows(all_sched_rows, monday_str)
        schedule     = {r["day_of_week"]: r["hours"] for r in active_sched if r["hours"] > 0}
        active_dist  = active_distribution_rows(
            [r for r in all_dist_rows if r["week_number"] == week_number], monday_str
        )
        tasks_for_week = [{"task_id": r["task_id"], "task_name": r["tasks"]["name"],
                           "task_color": r["tasks"].get("color"), "hours_per_week": r["hours_per_week"],
                           "schedule_rule": r["tasks"].get("schedule_rule"),
                           "is_fill": r["tasks"].get("is_fill", False),
                           "priority": r["tasks"].get("priority")}
                          for r in active_dist]
        preferred_days_wk = {r["task_id"]: r["preferred_days"] for r in active_dist if r.get("preferred_days")}
        task_map       = {t["task_id"]: t for t in tasks_for_week}

        # Use the full 5-day schedule so cross-month days get allocations too
        week_schedule = dict(schedule)

        if week_number >= from_week:
            allocations, _warnings = distribute_week(tasks_for_week, week_schedule, person["name"], preferred_days_wk)
        else:
            allocations = distribute_week_proportional(tasks_for_week, week_schedule)

        days       = []
        week_total = 0.0

        for dow in range(1, 6):
            actual_date    = monday + timedelta(days=dow - 1)
            is_other_month = actual_date.month != month

            scheduled_hrs = schedule.get(dow, 0.0)
            is_work_day   = scheduled_hrs > 0
            is_absent     = str(actual_date) in absent_dates

            # Week-row totals should match the visible Mon-Fri row, including
            # cross-month days that are rendered in the same week block.
            if is_work_day and not is_absent:
                week_total += scheduled_hrs

            daily_tasks = []
            if is_work_day and not is_absent:
                for task_id, hours in allocations.get(dow, {}).items():
                    if hours > 0 and task_id in task_map:
                        t = task_map[task_id]
                        daily_tasks.append({
                            "task_id":   task_id,
                            "task_name": t["task_name"],
                            "task_color": t["task_color"],
                            "hours":     hours,
                        })

            days.append({
                "date":            str(actual_date),
                "day_of_week":     dow,
                "day_name":        DAY_NAMES[dow - 1],
                "is_work_day":     is_work_day,
                "scheduled_hours": scheduled_hrs,
                "is_absent":       is_absent,
                "is_other_month":  is_other_month,
                "tasks":           daily_tasks,
            })

        weeks.append({
            "week_index":    week_index,
            "week_number":   week_number,
            "rules_applied": week_number >= from_week,
            "week_start":    str(monday),
            "total_hours":   week_total,
            "days":          days,
        })

    return {
        "person_id":    person_id,
        "person_name":  person["name"],
        "year":         year,
        "month":        month,
        "weekly_total": weekly_total,
        "weeks":        weeks,
    }
